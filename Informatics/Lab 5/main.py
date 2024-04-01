from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment
from string import ascii_uppercase
import utils

workbook = Workbook()
sheet = workbook.active


def get_col_name(n):
    if n <= 26:
        return ascii_uppercase[n - 1]
    return ascii_uppercase[n // 26 - 1] + ascii_uppercase[n % 26 - 1]


A = 7389
C = 17178

# A = 2187
# C = 30327

sheet["B1"] = "A ="
sheet["C1"] = A
sheet["B2"] = "C ="
sheet["C2"] = C


for row in range(4, 16):
    res = utils.calculateX(row - 3)
    sheet.cell(row=row, column=1).value = res[0]
    sheet.cell(row=row, column=2).value = res[1]
    sheet.cell(row=row, column=3).value = res[2]

for row in range(4, 16):
    sheet.cell(row=row, column=5).value = f"B{row - 3} ="
    if row >= 10:
        sheet.cell(row=row, column=6).value = f"-B{row - 3 - 6}="


for row in range(4, 16):
    column_cnt = 7
    for column in range(7, 26):
        sheet.column_dimensions[ascii_uppercase[column - 1]].width = 2

        if column in [11, 16, 21]:
            sheet.cell(row=row, column=column).value = "."
            continue

        cur_bit = 16 + 7 - column_cnt - 1
        BIT_LENGTH = f"_xlfn.FLOOR.MATH(LOG(ABS($C${row}), 2)) + 1"
        BINARY = f"MOD(_xlfn.BITRSHIFT(ABS($C${row}), {cur_bit}), 2)"
        NEG_BINARY = f"MOD(_xlfn.BITRSHIFT(ABS($C${row}) - 1, {cur_bit}), 2)"
        POS_REP = f"IF({cur_bit} < {BIT_LENGTH}, {BINARY}, 0)"
        NEG_REP = f"IF({cur_bit} < {BIT_LENGTH}, 1 - {NEG_BINARY}, 1)"
        REP = f"IF($C${row} > 0, {POS_REP}, {NEG_REP})"

        sheet.cell(row=row, column=column).value = f"={REP}"

        column_cnt += 1


"""
B1+B2, B2+B3, B2+B7, B7+B8, B8+B9, B1+B8, B11+B3
"""


def insert_res_n(res_n, row):
    sheet.cell(row=row, column=5).value = CellRichText(
        f"B{res_n}", TextBlock(InlineFont(vertAlign="subscript"), "(2)")
    )
    for column in range(7, 26):
        sheet.cell(
            row=row, column=column
        ).value = f"={ascii_uppercase[column - 1]}{res_n + 3}"


def insert_sum(line, n1, n2):
    FN_LINE = line
    SN_LINE = line + 1
    DASH_LINE = line + 2
    SUM_LINE = line + 3
    OF_LINE = line - 1

    insert_res_n(n1, FN_LINE)
    insert_res_n(n2, SN_LINE)

    sheet.cell(row=SN_LINE, column=4).value = "+"

    for col in range(5, 26):
        sheet.cell(row=DASH_LINE, column=col).value = "-" * (5 if col <= 6 else 2)

    for bit in range(25, 6, -1):
        if bit in [11, 16, 21]:
            sheet.cell(row=SUM_LINE, column=bit).value = "."
            continue
        shift = 2 if bit in [12, 17, 22] else 1

        SUM_F = f"{get_col_name(bit)}{FN_LINE} + {get_col_name(bit)}{SN_LINE} + {get_col_name(bit)}{OF_LINE}"

        BINARY_SUM = f"=MOD({SUM_F}, 2)"
        sheet.cell(row=SUM_LINE, column=bit).value = BINARY_SUM

        NEXT_OVERFLOW = f"=IF(SUM({get_col_name(bit)}{OF_LINE}:{get_col_name(bit)}{SN_LINE}) > 1, 1, 0)"
        sheet.cell(row=OF_LINE, column=bit - shift).value = NEXT_OVERFLOW

        CARRY_FLAG = f"=IF(F{OF_LINE} = 1, 1, 0)"
        sheet.cell(row=line + 5, column=7).value = "CF ="
        sheet.cell(row=line + 5, column=9).value = CARRY_FLAG

        PARITY_FLAG = f"=IF(MOD(SUM(Q{SUM_LINE}:Y{SUM_LINE}), 2) = 0, 1, 0)"
        sheet.cell(row=line + 5, column=11).value = "PF ="
        sheet.cell(row=line + 5, column=13).value = PARITY_FLAG

        AUX_CARRY_FLAG = f"=IF(T{OF_LINE} = 1, 1, 0)"
        sheet.cell(row=line + 5, column=15).value = "AF ="
        sheet.cell(row=line + 5, column=17).value = AUX_CARRY_FLAG

        ZERO_FLAG = f"=IF(SUM({get_col_name(6)}{SUM_LINE}:{get_col_name(25)}{SUM_LINE}) = 0, 1, 0)"
        sheet.cell(row=line + 6, column=7).value = "ZF ="
        sheet.cell(row=line + 6, column=9).value = ZERO_FLAG

        SIGN_FLAG = f"=IF({get_col_name(7)}{SUM_LINE} = 1, 1, 0)"
        sheet.cell(row=line + 6, column=11).value = "SF ="
        sheet.cell(row=line + 6, column=13).value = SIGN_FLAG

        OVERFLOW_FLAG = f"=IF(F{OF_LINE} <> G{OF_LINE}, 1, 0)"
        sheet.cell(row=line + 6, column=15).value = "OF ="
        sheet.cell(row=line + 6, column=17).value = OVERFLOW_FLAG

    MERGED = f'=SUBSTITUTE(_xlfn.TEXTJOIN("",1,G{SUM_LINE}:Y{SUM_LINE}), ".", "")'
    sheet.cell(row=SUM_LINE + 4, column=7).value = MERGED
    merged_add = get_col_name(7) + str(SUM_LINE + 4)
    POS_CONV = f"_xlfn.BITLSHIFT(BIN2DEC(LEFT(RIGHT({merged_add}, 15), 7)), 8) + BIN2DEC(RIGHT({merged_add}, 8))"
    INVERTED = f"SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({merged_add}, 0, 2), 1, 0), 2, 1)"
    NEG_CONV = f"(_xlfn.BITLSHIFT(BIN2DEC(LEFT(RIGHT({INVERTED}, 15), 7)), 8) + BIN2DEC(RIGHT({INVERTED}, 8)) + 1) * (-1)"
    CONV = f'=IF(LEFT({merged_add}, 1) = "0", {POS_CONV}, {NEG_CONV})'

    sheet.cell(row=SUM_LINE, column=27).value = CONV
    sheet.cell(row=SUM_LINE, column=26).value = CellRichText(
        TextBlock(InlineFont(vertAlign="subscript"), "(2)"), "="
    )
    sheet.cell(row=SUM_LINE, column=28).value = CellRichText(
        TextBlock(InlineFont(vertAlign="subscript"), "(10)")
    )

    sheet.cell(row=DASH_LINE - 1, column=33).value = CellRichText(
        f"X{n2}", TextBlock(InlineFont(vertAlign="subscript"), "(10)")
    )
    sheet.cell(row=DASH_LINE - 2, column=33).value = CellRichText(
        f"X{n1}", TextBlock(InlineFont(vertAlign="subscript"), "(10)")
    )
    sheet.cell(row=DASH_LINE - 1, column=32).value = "+"
    sheet.cell(row=DASH_LINE, column=31).value = "="
    sheet.cell(row=DASH_LINE, column=32).value = "-" * 5
    sheet.cell(row=DASH_LINE, column=33).value = "-" * 5
    sheet.cell(row=DASH_LINE, column=34).value = "-" * 5
    sheet.cell(row=FN_LINE, column=34).value = f"=C{3 + n1}"
    sheet.cell(row=SN_LINE, column=34).value = f"=C{3 + n2}"
    sheet.cell(
        row=SUM_LINE, column=34
    ).value = f"={get_col_name(34)}{FN_LINE} + {get_col_name(34)}{SN_LINE}"
    sheet.cell(row=SUM_LINE, column=35).value = CellRichText(
        TextBlock(InlineFont(vertAlign="subscript"), "(10)")
    )

    if1 = f'IF(AND(C{3 + n1} < 0, C{3 + n2} < 0), "При сложении двух отрицательных слагаемых", "При сложении положительного и отрицательного слагаемых")'
    if2 = f'IF(AND(C{3 + n1} >= 0, C{3 + n2} >= 0), "При сложении двух положительных слагаемых", {if1})'
    if3 = f'IF(AA{SUM_LINE} >= 0, "положительное число.", "отрицательное число.")'
    if4 = f'IF(Q{line + 6} = 1, "Зафиксировано переполнение, результат выполнения операции не соответствует сумме десятичных экивалентов", "Результат выполнения операции верный и корректный, соответствует сумме десятичных экивалентов.")'

    sheet.cell(
        row=SUM_LINE - 3, column=36
    ).value = f'=CONCATENATE({if2}, " получено ", {if3}, " ", {if4})'
    sheet.cell(row=SUM_LINE - 3, column=36).alignment = Alignment(wrapText=True)
    sheet.merge_cells(
        start_row=SUM_LINE - 3, start_column=36, end_row=SUM_LINE, end_column=40
    )


insert_sum(19, 1, 2)
insert_sum(29, 2, 3)
insert_sum(39, 2, 7)
insert_sum(49, 7, 8)
insert_sum(59, 8, 9)
insert_sum(69, 1, 8)
insert_sum(79, 11, 3)

workbook.save(filename="result.xlsx")
